from os import listdir
from os.path import isfile, join
import pandas as pd
import openpyxl
from time import clock
import random
import math

class Utils:
    def __init__(self):
        self.ext_txt = ".txt"
        self.ext_dat = ".dat"
        self.ext_xlsx = ".xlsx"

    def make_deep_cas9_input(self, path, dict_arr, init_arr, batch_size):
        pam_seq = init_arr[0]
        add_seq1_len = init_arr[1]
        spacer_len = init_arr[2]
        add_seq2_len = init_arr[3]
        pam_len = len(pam_seq)

        # del duplicates
        tmp_set = set()
        for data_dict in dict_arr:
            for chr_key, val_dict in data_dict.items():
                for trnscrpt_id, vals_arr in val_dict.items():
                    for trgt_idx in range(1, len(vals_arr)):
                        tmp_set.add(vals_arr[trgt_idx][0])

        tmp_list = list(tmp_set)
        tmp_chunks = [tmp_list[x:x + batch_size] for x in range(0, len(tmp_list), batch_size)]

        for idx in range(len(tmp_chunks)):
            with open(path + "_" + str(idx) + self.ext_txt, 'a') as f:
                f.write("Target number\t30 bp target sequence (4 bp + 20 bp protospacer + PAM + 3 bp)\n")
                for seq_str in tmp_chunks[idx]:
                    f.write(seq_str + "\n")

    def make_excel_after_sorting(self, path, result_list, init_arr):
        pam_seq = init_arr[0]
        add_seq1_len = init_arr[1]
        spacer_len = init_arr[2]
        add_seq2_len = init_arr[3]
        clvg_site = init_arr[4]
        pam_len = len(pam_seq)

        workbook = openpyxl.Workbook()
        sheet = workbook.active

        row = 1
        sheet.cell(row=row, column=1, value="index")
        sheet.cell(row=row, column=2, value='chromosome')
        sheet.cell(row=row, column=3, value='Target gene name')
        sheet.cell(row=row, column=4, value='Description')
        sheet.cell(row=row, column=5, value='Ensembl transcript ID')
        sheet.cell(row=row, column=6, value='Ensembl Gene ID')
        sheet.cell(row=row, column=7, value='Strand')
        sheet.cell(row=row, column=8, value='order sgRNA Target sequence')
        sheet.cell(row=row, column=9, value='order Target context sequence')
        sheet.cell(row=row, column=10, value='order PAM')
        sheet.cell(row=row, column=11, value='cleavage site')
        sheet.cell(row=row, column=12, value='DeepCas9 score')

        for val_arr in result_list:
            row += 1
            context_seq = val_arr[6]
            sheet.cell(row=row, column=1, value=str(row - 1))
            sheet.cell(row=row, column=2, value=val_arr[0])
            sheet.cell(row=row, column=3, value=val_arr[1])
            sheet.cell(row=row, column=4, value=val_arr[2])
            sheet.cell(row=row, column=5, value=val_arr[3])
            sheet.cell(row=row, column=6, value=val_arr[4])
            sheet.cell(row=row, column=7, value=val_arr[5])
            sheet.cell(row=row, column=8, value=context_seq[add_seq1_len:-add_seq2_len])
            sheet.cell(row=row, column=9, value=context_seq)
            sheet.cell(row=row, column=10, value=context_seq[add_seq1_len + spacer_len:-add_seq2_len])
            sheet.cell(row=row, column=11, value=val_arr[7])
            sheet.cell(row=row, column=12, value=val_arr[8])

        workbook.save(filename=path + self.ext_xlsx)

    def make_tsv_after_sorting(self, path, result_list, init_arr):
        pam_seq = init_arr[0]
        add_seq1_len = init_arr[1]
        spacer_len = init_arr[2]
        add_seq2_len = init_arr[3]
        clvg_site = init_arr[4]
        pam_len = len(pam_seq)

        with open(path + self.ext_txt, 'a') as f:
            row = 1
            f.write("index\tchromosome\tTarget gene name\tDescription\tEnsembl transcript ID\tEnsembl Gene ID\tStrand\torder sgRNA Target sequence\torder Target context sequence\torder PAM\tcleavage site\tDeepCas9 score\n")
            for val_arr in result_list:
                tmp_line = str(row) + "\t"
                for idx in range(len(val_arr)):
                    if idx == 6:
                        context_seq = val_arr[idx]
                        tmp_line += context_seq[add_seq1_len:-add_seq2_len] + "\t"
                        tmp_line += context_seq + "\t"
                        tmp_line += context_seq[add_seq1_len + spacer_len:-add_seq2_len] + "\t"
                    elif idx > 8:
                        pass
                    else:
                        tmp_line += str(val_arr[idx]) + "\t"
                f.write(tmp_line[:-1] + "\n")
                row += 1












